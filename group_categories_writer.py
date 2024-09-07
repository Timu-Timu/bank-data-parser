import openpyxl


class GroupCategoriesWriter:
    def __init__(self, excel_file: str):
        """Initialize the CategoryStore with an empty dictionary."""
        self.categories = {}
        self.new_samples = {}  # Dictionary to store new samples
        self.load_categories_from_excel(excel_file)

    def get_or_add_category(self, title: str) -> str:
        """
        Check if the category or title exists in the store.

        If it exists, return the category name. If not, prompt the user
        for a new category name, add the new pair, and return the category name.

        Args:
            title (str): The title associated with the category.

        Returns:
            str: The category name.
        """
        # Check if the title already exists
        # for stored_title, stored_category in self.categories.items():
        #     if stored_title == title:
        #         return stored_category
        if title in self.categories:
            return self.categories[title]

        # If it doesn't exist, check if it's a new sample
        if title in self.new_samples:
            return self.new_samples[title]

        # If neither exists, ask the user for a new category name
        new_category = input(f"Category for '{title}' do not exist. Please enter a new category name: ")

        # Add the new pair to the dictionary
        self.new_samples[title] = new_category
        return new_category

    def load_categories_from_excel(self, excel_file: str):
        """
        Load categories and titles from an Excel file into the categories dictionary.

        Args:
            excel_file (str): The path to the Excel file to read from.
        """
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Iterate through the rows in the sheet, skipping the header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            title, category = row  # Unpack the values from the row
            if category and title:  # Ensure both values are not None
                self.categories[title] = category  # Add to the categories dictionary

    def write_new_samples_to_excel(self, excel_file: str):
        """
        Write new samples (title and category) to the existing Excel file.

        Args:
            excel_file (str): The path to the Excel file to write to.
        """
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Find the next available row to write the new samples
        next_row = sheet.max_row + 1

        # Iterate through the new_samples dictionary and write to the sheet
        for title, category in self.new_samples.items():
            sheet.cell(row=next_row, column=1, value=title)
            sheet.cell(row=next_row, column=2, value=category)
            next_row += 1

        # Save the workbook
        workbook.save(excel_file)
