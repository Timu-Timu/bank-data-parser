from datetime import datetime

import openpyxl

class ResultsWriter:
    def __init__(self):
        """Initialize the ResultsWriter with an empty list to store parsed data."""
        self.parsed_data = []

    def write_sample_info(self, title: str, amount: float, date: str, category_writer):
        """
        Store the parsed span information (category, title, amount).

        Args:
            title (str): The title.
            amount (float): The amount.
            date (str): date of operation
            category_writer: Writer service
        """
        is_valid = self.sample_validation(title, amount, date)
        if not is_valid:
            return
        equalizer_category = category_writer.get_or_add_category(title)

        self.parsed_data.append((equalizer_category, title, amount, date))

    def sample_validation(self, title: str, amount: float, date: str):
        if "тимур владимирович а" in title.lower():
            return False
        else:
            return True

    def write_results_to_excel(self):
        """Write the parsed data to a new Excel file."""
        # Get the current date and time
        now = datetime.now()
        file_name = f"output/export_{now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        # Create a new workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write the header row
        sheet['A1'] = 'Категория'
        sheet['B1'] = 'Наименование'
        sheet['C1'] = 'Цена'
        sheet['D1'] = 'Дата'
        sheet['E1'] = 'Со счета'

        # Find the next available row to write the parsed data
        next_row = 2

        # Iterate through the parsed_data list and write to the sheet
        for category, title, amount, date in self.parsed_data:
            sheet.cell(row=next_row, column=1, value=category)
            sheet.cell(row=next_row, column=2, value=title)
            sheet.cell(row=next_row, column=3, value=amount)
            sheet.cell(row=next_row, column=4, value=date)
            sheet.cell(row=next_row, column=5, value="Дебетовые карты БКС RUB")
            next_row += 1

        # Save the workbook to the new file
        workbook.save(file_name)
        print(f"Results written to {file_name}")

